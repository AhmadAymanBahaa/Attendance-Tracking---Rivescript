import openpyxl

workbookName = 'attendance.xlsx'
workbook = openpyxl. load_workbook(workbookName)


def export(name, id, time):

    for wks in workbook.worksheets:
        wks.cell(1,1).value = "Name"
        wks.cell(1,2).value = "ID"
        wks.cell(1,3).value = "Time"

        x = wks.max_row + 1
        wks.cell(x, 1 ).value = name
        wks.cell(x, 2).value = id
        wks.cell(x, 3).value = time

    workbook.save(workbookName)
    workbook.close()

